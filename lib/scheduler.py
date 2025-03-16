import json
import random
from datetime import datetime, timedelta
from collections import defaultdict
import openpyxl

class GradeTimetableGenerator:
    def __init__(self):
        self.load_settings()
        self.main_subjects = ["语文", "数学", "英语"]
        self.init_time_slots()
        self.teacher_pool = self.create_teacher_pool()
        # 修复数据结构为双层字典
        self.grade_history = defaultdict(lambda: defaultdict(list))

    def load_settings(self):
        with open('settings.json', 'r', encoding='utf-8') as f:
            self.settings = json.load(f)
        
        # 基础设置
        self.grade = self.settings['basic'].get('grade', '一')
        self.class_count = int(self.settings['basic'].get('class_count', 6))
        self.days = int(self.settings['basic']['days'])
        self.max_duration = int(self.settings['basic']['max_duration'])
        
        # 解析时间规则
        self.time_rules = self.parse_time_rules()  # 新增关键代码
        
        # 解析科目配置
        self.subjects = []
        for item in self.settings['subjects']:
            if isinstance(item, dict):
                self.subjects.append({
                    'name': item['name'],
                    'teachers': item.get('teachers', 0)
                })
            else:
                self.subjects.append({'name': item, 'teachers': 0})
    
    def parse_time_rules(self):
        """解析时间规则"""
        day_map = {"周一":0, "周二":1, "周三":2, "周四":3, "周五":4, "周六":5, "周日":6}
        rules = {}
        for rule in self.settings['time_rules']:
            if rule[1] == "有":
                day = day_map[rule[0]]
                start = datetime.strptime(rule[2], "%H:%M")
                end = datetime.strptime(rule[3], "%H:%M")
                delta = (end - start).seconds // 60
                num_slots = delta // self.max_duration
                rules[day] = {
                    "start": start,
                    "slots": [start + timedelta(minutes=i*self.max_duration) 
                             for i in range(num_slots)]
                }
        return rules
    
    def init_time_slots(self):
        """初始化时间槽"""
        self.time_slots = {}
        for day in range(self.days):
            if day in self.time_rules:  # 现在self.time_rules已正确定义
                slots = self.time_rules[day]['slots']
                self.time_slots[day] = {
                    "morning": [t for t in slots if t.hour < 12],
                    "afternoon": [t for t in slots if t.hour >= 12]
                }

    def create_teacher_pool(self):
        pool = defaultdict(list)
        for subj in self.subjects:
            if subj['teachers'] > 0:
                pool[subj['name']] = [
                    f"{subj['name']}教师{i+1}"
                    for i in range(subj['teachers'])
                ]
        return pool
    
    def generate_grade(self):
        grade_timetable = {}
        for class_id in range(1, self.class_count+1):
            class_name = f"{self.grade}年级{class_id}班"
            grade_timetable[class_name] = self.generate_class_timetable(class_id)
        return grade_timetable
    
    def generate_class_timetable(self, class_id):
        """生成单个班级课表"""
        timetable = {}
        for day in range(self.days):
            day_name = f"周{['一','二','三','四','五','六','日'][day]}"
            timetable[day] = {"day": day_name, "courses": []}
            
            morning_slots = self.time_slots.get(day, {}).get('morning', [])
            afternoon_slots = self.time_slots.get(day, {}).get('afternoon', [])
            
            main_ratio = 0.7 if day < 4 else 0.4
            morning_main_ratio = 0.8
            
            # 生成上午课程
            for slot in morning_slots:
                course = self.select_course(class_id, day, slot, is_morning=True,
                                           main_ratio=morning_main_ratio)
                entry = {"time": slot.strftime("%H:%M"), "course": course}
                timetable[day]["courses"].append(entry)
                # 新增：记录到历史
                self.grade_history[class_id][day].append(entry)
            
            # 生成下午课程
            for slot in afternoon_slots:
                course = self.select_course(class_id, day, slot, is_morning=False,
                                          main_ratio=main_ratio)
                entry = {"time": slot.strftime("%H:%M"), "course": course}
                timetable[day]["courses"].append(entry)
                # 新增：记录到历史
                self.grade_history[class_id][day].append(entry)
        
        return timetable

    def select_course(self, class_id, day, slot, is_morning, main_ratio):
        """课程选择逻辑"""
        """修复后的课程选择逻辑"""
        prev_day = (day - 1) % self.days
        time_key = slot.strftime("%H:%M")
        
        # 修正历史记录访问方式
        class_history = [
            c["course"] 
            for c in self.grade_history[class_id][prev_day]
            if c["time"] == time_key
        ]
        # 全校教师资源约束
        same_time_courses = []
        current_time = slot.strftime("%H:%M")
        for class_data in self.grade_history.values():  # 每个班级的数据
        # 获取该班级当天的课程安排
            day_courses = class_data.get(day, [])
        # 查找相同时间的课程
            for course in day_courses:
                if course["time"] == current_time:
                    same_time_courses.append(course["course"])
        
        candidates = []
        for subj_config in self.subjects:
            subject = subj_config['name'] if isinstance(subj_config, dict) else subj_config
            is_main = subject in self.main_subjects
            
            # 检查教师资源
            teacher_available = self.check_teacher_availability(subject, same_time_courses)
            
            if teacher_available and self.check_constraints(subject, is_morning, main_ratio):
                candidates.append(subject)
        
        # 新增：检查当前班级当天已排课程
        current_day_history = self.grade_history[class_id][day]
        if current_day_history:
            last_course = current_day_history[-1]["course"]
            candidates = [c for c in candidates if c != last_course]
        
        # 排除历史冲突
        candidates = [c for c in candidates if c not in class_history]
        
        if candidates:
            selected = random.choice(candidates)
            # 记录教师使用情况
            self.record_teacher_usage(selected, slot)
            return selected
        return "自习"

    def check_teacher_availability(self, subject, same_time_courses):
        max_parallel = len(self.teacher_pool.get(subject, []))
        current_used = same_time_courses.count(subject)
        
        # 处理无教师配置科目
        if max_parallel == 0 and any(s['name'] == subject for s in self.subjects):
            return True
        return current_used < max_parallel
    
    def record_teacher_usage(self, subject, slot):
        """记录教师使用情况（虚拟实现）"""
        # 实际使用时可根据具体时间分配教师
        pass

    def check_constraints(self, subject, is_morning, main_ratio):
        """检查排课约束"""
        if is_morning:
            if subject in self.main_subjects:
                return random.random() < main_ratio
            return random.random() > main_ratio
        else:
            return True

    def export_to_excel(self, timetable, filename="grade_timetable.xlsx"):
        """增强版Excel导出方法"""
        import os
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        try:
            # 验证模板路径
            template_path = os.path.join("src/tps", "c12.xlsx")
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"模板文件不存在于：{os.path.abspath(template_path)}")

            # 加载模板
            wb = load_workbook(template_path)
            template_ws = wb["Sheet1"]

            # 列映射配置
            column_mapping = {0:2, 1:3, 2:4, 3:5, 4:6}  # 周一至周五对应C-G列

            # 处理每个班级
            for class_idx, (class_name, schedule) in enumerate(timetable.items(), 1):
                print(f"正在生成 {class_name} ({class_idx}/{len(timetable)})")
                
                # 复制模板工作表
                new_ws = wb.copy_worksheet(template_ws)
                new_ws.title = class_name[:25]
                
                # 更新标题
                new_ws['B1'] = f"{class_name}课程表"
                
                # 收集所有时间段
                all_times = sorted({
                    c["time"] for day in schedule.values()
                    for c in day["courses"]
                }, key=lambda x: datetime.strptime(x, "%H:%M"))
                
                # 填充数据（从第3行开始）
                for row_idx, time in enumerate(all_times, 3):
                    # 写入时间列
                    #new_ws.cell(row=row_idx, column=2, value=time)
                    
                    # 填充每日课程
                    for day_num in range(self.days):
                        courses = [
                            c["course"] for c in schedule[day_num]["courses"]
                            if c["time"] == time
                        ]
                        col = column_mapping[day_num]
                        cell = new_ws.cell(row=row_idx, column=col)
                        cell.value = "\n".join(courses)
                        cell.alignment = Alignment(wrapText=True, vertical='top')

            # 删除模板页并保存
            wb.remove(template_ws)
            wb.save(filename)
            print(f"成功生成：{os.path.abspath(filename)}")
            return True

        except Exception as e:
            print(f"生成失败：{str(e)}")
            return False
        
        
if __name__ == "__main__":
    generator = GradeTimetableGenerator()
    grade_timetable = generator.generate_grade()
    
    # 打印示例
    for class_name, timetable in grade_timetable.items():
        #print(f"\n{class_name}")
        for day in timetable.values():
            #print(f"\n{day['day']}")
            for course in day["courses"]:
                #print(f"{course['time']} - {course['course']}")
                pass
    
    generator.export_to_excel(grade_timetable, "grade_timetable.xlsx")